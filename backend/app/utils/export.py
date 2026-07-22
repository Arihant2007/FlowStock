import io
from datetime import datetime
from typing import Literal

import pandas as pd
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from fpdf import FPDF

ExportFormat = Literal["csv", "excel", "pdf"]


def generate_export(
    df: pd.DataFrame, report_name: str, export_format: ExportFormat = "excel", metadata: dict | None = None
) -> StreamingResponse:
    """Generate a StreamingResponse containing either a CSV or an Excel file.

    If Excel, applies basic formatting:
      - Bold headers
      - Auto-fit column widths
      - Freeze top row
    """
    date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{report_name}_{date_str}"

    if export_format == "csv":
        # CSV Injection protection
        for col in df.select_dtypes(include=['object']):
            df[col] = df[col].apply(
                lambda x: f"'{x}" if isinstance(x, str) and str(x).startswith(('=', '+', '-', '@')) else x
            )

        stream = io.StringIO()
        df.to_csv(stream, index=False)
        stream.seek(0)

        return StreamingResponse(
            iter([stream.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
        )

    # Excel formatting
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Report")
        worksheet = writer.sheets["Report"]

        # Freeze top row
        worksheet.freeze_panes = "A2"

        # Format headers and auto-fit columns
        for col_idx, col in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)

            # Header formatting
            header_cell = worksheet[f"{col_letter}1"]
            header_cell.font = Font(bold=True)
            header_cell.alignment = Alignment(horizontal="center")

            # Calculate max width for auto-fit
            max_len = len(str(col))
            for row in range(2, len(df) + 2):
                cell_val = worksheet[f"{col_letter}{row}"].value
                if cell_val is not None:
                    max_len = max(max_len, len(str(cell_val)))

            # Apply width with padding
            worksheet.column_dimensions[col_letter].width = max_len + 2

    output.seek(0)

    if export_format == "pdf":
        class PDF(FPDF):
            def header(self):
                self.set_font("helvetica", "B", 14)
                self.cell(0, 10, "ITC FMCG Warehouse Management System", ln=True, align="C")
                self.set_font("helvetica", "B", 12)
                self.cell(0, 10, f"Report: {report_name.replace('_', ' ')}", ln=True, align="C")
                self.set_font("helvetica", "", 10)
                self.cell(0, 10, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align="C")
                if metadata:
                    for k, v in metadata.items():
                        self.cell(0, 6, f"{k}: {v}", ln=True, align="C")
                self.ln(5)

            def footer(self):
                self.set_y(-15)
                self.set_font("helvetica", "I", 8)
                self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")

        # Use landscape for wide tables
        orientation = "L" if len(df.columns) > 5 else "P"
        pdf = PDF(orientation=orientation, format="A4")
        pdf.add_page()
        pdf.set_font("helvetica", size=8)

        # Calculate column widths
        page_width = pdf.w - 2 * pdf.l_margin
        col_count = len(df.columns)
        col_width = page_width / col_count if col_count > 0 else 0

        # Header
        pdf.set_font("helvetica", "B", 8)
        for col in df.columns:
            pdf.cell(col_width, 8, str(col), border=1, align="C")
        pdf.ln()

        # Rows
        pdf.set_font("helvetica", "", 8)
        for _, row in df.iterrows():
            for val in row:
                pdf.cell(col_width, 8, str(val)[:50], border=1) # Truncate long strings
            pdf.ln()

        pdf_bytes = pdf.output()
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}.pdf"},
        )

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )
